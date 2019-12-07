# Imports
import numpy as np
import scipy as sp
from scipy import linalg
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook


# Main
def main():
    # Gathers information, creates the admittance matrix, and creates the output file
    tolerance, input_path, output_path = ask_user()
    values, input_line, bus_size, tot_implicit, input_bus = initialize(input_path)
    g_matrix, b_matrix = admittance(input_line)
    wb, ws1, ws2, ws3 = create_file()

    # Creates the first iteration of the mismatch
    known_power = track_type(input_bus, bus_size)
    converged, mismatches, max_p, max_q, p_bus, q_bus = mismatch(tolerance, g_matrix, b_matrix, values, bus_size, tot_implicit)
    mismatches_old = mismatches.copy()
    diverged = False
    iteration = 0
    write_cont(ws2, iteration, max_p, max_q, p_bus, q_bus)

    # While not within tolerance but not diverging
    while not converged and not diverged:
        # Creates the Jacobian and solves for this iteration's new voltage and angle values
        j_matrix = build_jacobian(values, g_matrix, b_matrix, known_power, mismatch_size=len(mismatches))
        corrections = correct(j_matrix, mismatches)
        values = update(corrections, values, known_power)

        # Creates the next iteration's mismatch equations and checks for convergence / divergence
        converged, mismatches, max_p, max_q, p_bus, q_bus = mismatch(tolerance, g_matrix, b_matrix, values, bus_size, tot_implicit)
        for i in range(len(mismatches)):
            if mismatches[i] >= mismatches_old[i]:
                diverged = True
        iteration += 1
        write_cont(ws2, iteration, max_p, max_q, p_bus, q_bus)
        mismatches_old = mismatches.copy()

    wb.save(output_path + '\\output_results.xlsx')
    if diverged:
        print("No solution found.")
    else:
        print("Solution found!")
        # TODO print output files


# Asks the user for tolerance, import file path, and output file path
def ask_user():
    print("What is your desired tolerance (in MVA)?")
    print("Eg. 0.1")
    #a = float(input()) / 100.0
    a = 0.1 / 100.0  # TODO: remove
    print("What is your input file path?")
    print("Eg. C:\\Users\\jtsch\\Desktop\\system_SampleInput.xlsx")
    #b = input()
    b = "C:\\Users\\jtsch\\Desktop\\system_SampleInput.xlsx"  # TODO: remove
    print("What is your output file path?")
    print("Eg. C:\\Users\\jtsch\\Desktop")
    #c = input()
    c = "C:\\Users\\jtsch\\Desktop"  # TODO: remove
    return a, b, c


# Initializes value matrix using the input file
def initialize(input_path):
    wb = load_workbook(input_path)
    input_bus = wb['BusData']
    input_line = wb['LineData']
    bus_size = input_bus.max_row - 1
    values = np.zeros((bus_size, 5))
    tot_implicit = 0
    for i in range(bus_size): # p_inj, q_inj, v, theta, type
        values[i, 0] = float(input_bus.cell(row=(i+2), column=5).value)-float(input_bus.cell(row=(i+2), column=2).value)
        values[i, 0] /= 100.0
        values[i, 1] = -1*float(input_bus.cell(row=(i+2), column=3).value) / 100.0
        values[i, 2] = float(input_bus.cell(row=(i+2), column=6).value)
        values[i, 3] = 0.0
        bus_type = input_bus.cell(row=(i+2), column=4).value
        if bus_type == 'S':
            values[i, 4] = 0.0  # Slack bus
        elif bus_type == 'G':
            values[i, 4] = 1.0  # PV bus
            tot_implicit += 1
        else:
            values[i, 4] = 2.0  # PQ bus
            tot_implicit += 2
    return values, input_line, bus_size, tot_implicit, input_bus


# Admittance matrix Y = G + jB
def admittance(input_line):
    line_size = input_line.max_row - 1
    y = 1j * np.zeros((line_size, line_size))
    b_temp = np.zeros(line_size)

    # Populate admittance matrix indices ji, stores shunt admittance
    for i in range(input_line.max_row - 1):
        line_from = int(input_line.cell(row=(i+2), column=1).value)  # Line starts at this bus
        line_to = int(input_line.cell(row=(i+2), column=2).value)  # Line ends at this bus
        line_r = float(input_line.cell(row=(i+2), column=3).value)  # Line resistance
        line_x = float(input_line.cell(row=(i+2), column=4).value)  # Line reactance
        y[line_from - 1][line_to - 1] = -(line_r + 1j * line_x) ** -1  # Yji
        y[line_to - 1][line_from - 1] = -(line_r + 1j * line_x) ** -1  # Yij
        b_temp[line_from - 1] += float(input_line.cell(row=(i+2), column=5).value)  # Shunt admittance for start
        b_temp[line_to - 1] += float(input_line.cell(row=(i+2), column=5).value)  # Shunt admittance for end

    # Populate admittance matrix indices ii
    for i in range(line_size):
        y[i][i] = -sum([y[i][x] for x in range(line_size)]) + 0.5j * b_temp[i]
    g_matrix = y.real
    b_matrix = y.imag
    return g_matrix, b_matrix


# Creates the output file and its headers
def create_file():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "BusResults"
    ws1.cell(row=1, column=1).value = "Bus Number"
    ws1.cell(row=1, column=2).value = "V(p.u.)"
    ws1.cell(row=1, column=3).value = "Angle(deg)"
    ws1.cell(row=1, column=4).value = "Pinj. (MW)"
    ws1.cell(row=1, column=5).value = "Qinj. (MVAr)"
    ws1.cell(row=1, column=6).value = "Voltage Violation"
    ws2 = wb.create_sheet(title="ConvergenceHistory")
    ws2.cell(row=1, column=1).value = "Iteration"
    ws2.cell(row=1, column=2).value = "Max P Mismatch"
    ws2.cell(row=1, column=3).value = "P Bus"
    ws2.cell(row=1, column=4).value = "Max Q Mismatch"
    ws2.cell(row=1, column=5).value = "Q Bus"
    ws3 = wb.create_sheet(title="LineFlow")
    ws3.cell(row=1, column=1).value = "From Bus"
    ws3.cell(row=1, column=2).value = "To Bus"
    ws3.cell(row=1, column=3).value = "P Flow(MW)"
    ws3.cell(row=1, column=4).value = "Q Flow(MVAr)"
    ws3.cell(row=1, column=5).value = "S Flow(MVA)"
    ws3.cell(row=1, column=6).value = "MVA Violation"
    return wb, ws1, ws2, ws3


# Helps keep track of which powers are given and which need to be solved for explicitly
# 0 means the value is not known, else it is 1
def track_type(input_bus, bus_size):
    known_power = np.zeros((2, bus_size))  # 0 for unknown (explicit), 1 for known (implicit)
    for i in range(bus_size):
        bus_type = input_bus.cell(row=i+2, column=4).value
        if bus_type == "S":
            known_power[0][i], known_power[1][i] = 0, 0
        elif bus_type == "G":
            known_power[0][i], known_power[1][i] = 1, 0
        else:
            known_power[0][i], known_power[1][i] = 1, 1
    return known_power


# Calculates the mismatch, decides if converged, and which buses have largest mismatch for real and reactive power
def mismatch(tolerance, g_matrix, b_matrix, values, bus_size, tot_implicit):
    mismatches = np.zeros(tot_implicit)
    max_p = 0.0000
    max_q = 0.0000
    p_bus = -1
    q_bus = -1
    converged = True
    max_p, p_bus, converged, mismatches, count = write_p(tolerance, g_matrix, b_matrix, values, bus_size, max_p, p_bus, converged, mismatches)
    max_q, q_bus, converged, mismatches = write_q(tolerance, g_matrix, b_matrix, values, bus_size, max_q, q_bus, converged, mismatches, count)
    return converged, mismatches, max_p, max_q, p_bus, q_bus


# Creates the mismatch values for the real power equations
def write_p(tolerance, g_matrix, b_matrix, values, bus_size, max_p, p_bus, converged, mismatches):
    count = 0
    for k in range(bus_size):  # Calculate real power mismatch
        if values[k][4] == 1.0 or 2.0:  # PV or PQ bus, respectively
            p_sum = 0.0
            for i in range(bus_size):
                temp_var = g_matrix[k][i]*np.cos(values[k][3]-values[i][3]) + b_matrix[k][i]*np.sin(values[k][3]-values[i][3])
                p_sum += values[k][2]*values[i][2]*temp_var
            p_sum -= values[k][0]  # Subtracts the actual real power injected from the summation
            mismatches[count] = p_sum
            if p_sum > tolerance:
                converged = False
            if max_p < p_sum:
                max_p = p_sum
                p_bus = k + 1
            count += 1
    return max_p, p_bus, converged, mismatches, count


# Creates the mismatch values for the reactive power equations
def write_q(tolerance, g_matrix, b_matrix, values, bus_size, max_q, q_bus, converged, mismatches, count):
    for k in range(bus_size):  # Calculate reactive power mismatch
        if values[k][4] == 2.0:  # PQ bus
            q_sum = 0.0
            for i in range(bus_size):
                temp_var = g_matrix[k][i]*np.sin(values[k][3]-values[i][3]) - b_matrix[k][i]*np.cos(values[k][3]-values[i][3])
                q_sum += values[k][2]*values[i][2]*temp_var
                print(q_sum)
            q_sum -= values[k][1]  # Subtracts the actual reactive power injected from the summation
            #mismatches[count] = q_sum
            print(q_sum)
            print()
            if q_sum > tolerance:
                converged = False
            if max_q < q_sum:
                max_q = q_sum
                q_bus = k + 1
            count += 1
    return max_q, q_bus, converged, mismatches

'''
# Writes the history of this iteration to the output file
def write_cont(ws2, iteration, max_p, max_q, p_bus, q_bus):
    ws2.cell(row=(iteration + 2), column=1).value = iteration
    ws2.cell(row=(iteration + 2), column=2).value = max_p
    ws2.cell(row=(iteration + 2), column=3).value = p_bus
    ws2.cell(row=(iteration + 2), column=4).value = max_q
    ws2.cell(row=(iteration + 2), column=5).value = q_bus


# Creates the Jacobian matrix for the current iteration
def build_jacobian(values, g_matrix, b_matrix, known_power, mismatch_size):
    pass
    j_matrix = np.zeros((mismatch_size, mismatch_size))
    for i in range(mismatch_size):
        for j in range(mismatch_size):
            i_bus = mis
            j_bus =
            if i == j:
                if i_bus == 0 and j_pq == 0:
                    j_matrix = -values[1][i]-b_matrix[i][i]*values[2][i]**2
                elif i_pq == 0 and j_pq == 1:
                    j_matrix =
                elif i_pq == 1 and j_pq == 0:
                    j_matrix =
                else:
                    j_matrix =
            else:
                if i_pq == 0 and j_pq == 0:
                    j_matrix =
                elif i_pq == 0 and j_pq == 1:
                    j_matrix =
                elif i_pq == 1 and j_pq == 0:
                    j_matrix =
                else:
                    j_matrix =
    return j_matrix


# Creates the corrections needed for this iteration
def correct(j_matrix, mismatches):
    inv_j = -1 * np.linalg.inv(j_matrix)
    corrections = np.matmul(inv_j, mismatches)
    return corrections


# Updates the values matrix for this iteration
# Adds the angles first based on the real power known, then adds the voltages based on the reactive power known
def update(corrections, values, known_power):
    tracker = 0
    for i in range(2):
        for j in range(len(known_power)):
            if known_power[i][j] == 1:
                values[j][3] = corrections[tracker]
                tracker += 1
        for k in range(len(known_power)):
            if known_power[i][k] == 1:
                values[k][2] = corrections[tracker]
                tracker += 1
    return values'''


main()
