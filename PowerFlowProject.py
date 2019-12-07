# Imports
import numpy as np
import scipy as sp
from scipy import linalg
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook


s_base = 100.0


# Main
def main():
    iteration = 0
    tolerance, input_path, output_path = ask_user()
    wb, ws1, ws2, ws3, ws4, ws5, input_bus, input_line, bus_size, line_size, input_bus_file = make_output(input_path)
    wb.save(output_path + '\\output_results.xlsx')
    bus_info, load_info, load_size = initialize(input_bus, bus_size, input_bus_file)
    g_matrix, b_matrix = admittance(input_line, line_size, ws4, ws5)
    mismatch, max_mismatch = calc_mismatch(g_matrix, b_matrix, bus_info, load_info, bus_size, load_size, ws2, iteration)
    wb.save(output_path + '\\output_results.xlsx')
    jacobian = calc_jacobian(bus_info, load_info, g_matrix, b_matrix, bus_size, load_size)
    print()


# Asks the user for tolerance, import file path, and output file path
# a = user tolerance in pu
# b = user input file location
# c = user output location
def ask_user():
    print("What is your desired tolerance (in MVA)?")
    print("Eg. 0.1")
    #a = float(input()) / s_base
    a = 0.1 / s_base  # TODO: remove
    print("What is your input file path?")
    print("Eg. C:\\Users\\jtsch\\Desktop\\system_SampleInput.xlsx")
    #b = input()
    b = "C:\\Users\\jtsch\\Desktop\\system_SampleInput.xlsx"  # TODO: remove
    print("What is your output file path?")
    print("Eg. C:\\Users\\jtsch\\Desktop")
    #c = input()
    c = "C:\\Users\\jtsch\\Desktop"  # TODO: remove
    return a, b, c


# Creates the output file
# wb = output file
# ws1 = Bus Results
# ws2 = Convergence History
# ws3 = Line Flow
# ws4 = G Matrix
# ws5 = B Matrix
# input_bus = bus information put into a matrix; p_load, q_load, p_gen, v
# input_line = line information put into a matrix; from, to, r, x, b, rate
# bus_size = number of buses
# line_size = number of lines
# input_bus_file = input bus sheet file
def make_output(input_path):
    wb = load_workbook(input_path)
    input_bus_file = wb['BusData']
    input_line_file = wb['LineData']
    ws1 = wb.active

    ws1.title = "BusResults"
    ws1.cell(row=1, column=1).value = "Bus Number"
    ws1.cell(row=1, column=2).value = "V(p.u.)"
    ws1.cell(row=1, column=3).value = "Angle(deg)"
    ws1.cell(row=1, column=4).value = "Pinj. (MW)"
    ws1.cell(row=1, column=5).value = "Qinj. (MVAr)"
    ws1.cell(row=1, column=6).value = "Voltage Violation"

    ws2 = wb.create_sheet(title="ConvergenceHistory")
    ws2.cell(row=1, column=1).value = "Iteration"
    ws2.cell(row=1, column=2).value = "Max P Mismatch"
    ws2.cell(row=1, column=3).value = "P Bus"
    ws2.cell(row=1, column=4).value = "Max Q Mismatch"
    ws2.cell(row=1, column=5).value = "Q Bus"

    ws3 = wb.create_sheet(title="LineFlow")
    ws3.cell(row=1, column=1).value = "From Bus"
    ws3.cell(row=1, column=2).value = "To Bus"
    ws3.cell(row=1, column=3).value = "P Flow(MW)"
    ws3.cell(row=1, column=4).value = "Q Flow(MVAr)"
    ws3.cell(row=1, column=5).value = "S Flow(MVA)"
    ws3.cell(row=1, column=6).value = "MVA Violation"

    ws4 = wb.create_sheet(title="G Matrix")

    ws5 = wb.create_sheet(title="B Matrix")

    bus_size = input_bus_file.max_row - 1
    input_bus = np.zeros((bus_size, 4))
    for i in range(bus_size):
        input_bus[i, 0] = -1 * float(input_bus_file.cell(row=(i + 2), column=2).value) / s_base
        input_bus[i, 1] = -1 * float(input_bus_file.cell(row=(i + 2), column=3).value) / s_base
        input_bus[i, 2] = float(input_bus_file.cell(row=(i + 2), column=5).value) / s_base
        input_bus[i, 3] = float(input_bus_file.cell(row=(i + 2), column=6).value)

    line_size = input_bus_file.max_row - 1
    input_line = np.zeros((line_size, 6))
    for i in range(line_size):
        input_line[i, 0] = int(input_line_file.cell(row=(i + 2), column=1).value)
        input_line[i, 1] = int(input_line_file.cell(row=(i + 2), column=2).value)
        input_line[i, 2] = float(input_line_file.cell(row=(i + 2), column=3).value)
        input_line[i, 3] = float(input_line_file.cell(row=(i + 2), column=4).value)
        input_line[i, 4] = float(input_line_file.cell(row=(i + 2), column=5).value)
        input_line[i, 5] = float(input_line_file.cell(row=(i + 2), column=6).value)

    return wb, ws1, ws2, ws3, ws4, ws5, input_bus, input_line, bus_size, line_size, input_bus_file


# Admittance matrix Y = G + jB
def admittance(input_line, line_size, ws4, ws5):
    y = 1j * np.zeros((line_size, line_size))
    b_temp = np.zeros(line_size)

    # Populate admittance matrix indices ji, stores shunt admittance
    for i in range(line_size):
        line_from = int(input_line[i, 0])  # Line starts at this bus
        line_to = int(input_line[i, 1])  # Line ends at this bus
        line_r = input_line[i, 2]  # Line resistance
        line_x = input_line[i, 3]  # Line reactance
        y[line_from - 1][line_to - 1] = -(line_r + 1j * line_x) ** -1  # Yji
        y[line_to - 1][line_from - 1] = -(line_r + 1j * line_x) ** -1  # Yij
        b_temp[line_from - 1] += input_line[i, 4]  # Shunt admittance for start
        b_temp[line_to - 1] += input_line[i, 4]  # Shunt admittance for end

    # Populate admittance matrix indices ii
    for i in range(line_size):
        y[i][i] = -sum([y[i][x] for x in range(line_size)]) + 0.5j * b_temp[i]
    g_matrix = y.real
    b_matrix = y.imag
    for i in range(line_size):
        for j in range(line_size):
            ws4.cell(row=1+i,column=1+j).value = g_matrix[i, j]
            ws5.cell(row=1+i,column=1+j).value = b_matrix[i, j]
    return g_matrix, b_matrix


# Fill out matrices used to keep track of changes during the iterations
# p_load, q_load, p_inj, q_inj, v, angle
# bus_num, q_load, v, angle
# load_size = number of PQ buses
def initialize(input_bus, bus_size, input_bus_file):
    bus_info = np.zeros((bus_size, 6))
    load_info = np.zeros((bus_size, 4))
    load_size = 0
    for i in range(bus_size):
        bus_info[i, 0] = input_bus[i, 0]
        bus_info[i, 1] = input_bus[i, 1]
        bus_info[i, 2] = input_bus[i, 2] + input_bus[i, 0]
        bus_info[i, 3] = input_bus[i, 1] * -1
        bus_info[i, 4] = input_bus[i, 3]
        bus_type = input_bus_file.cell(row=2+i,column=4).value
        if bus_type == "D":
            load_info[load_size, 0] = i
            load_info[load_size, 1] = input_bus[i, 1] * -1
            load_info[load_size, 2] = input_bus[i, 3]
            load_size += 1
    return bus_info, load_info, load_size


# Creates the mismatch matrix for this iteration
# mismatch = all mismatch values stored
# max_mismatch = max p and q mismatch along with which bus
def calc_mismatch(g_matrix, b_matrix, bus_info, load_info, bus_size, load_size, ws2, iteration):
    mismatch_size = (bus_size - 1 + load_size)
    mismatch = np.zeros(mismatch_size)
    max_mismatch = np.zeros((2,2))
    count = 0
    for i in range(bus_size):
        if i>0:
            for j in range(bus_size):
                mismatch[count] += bus_info[i,4]*bus_info[j,4]*(g_matrix[i,j]*np.cos(bus_info[i,5]-bus_info[j,5])+b_matrix[i,j]*np.sin(bus_info[i,5]-bus_info[j,5]))
            mismatch[count] -= bus_info[i,2]
            if abs(mismatch[count]) > abs(max_mismatch[0,0]):
                max_mismatch[0, 0] = abs(mismatch[count])
                max_mismatch[0, 1] = i + 1
            count += 1
    for i in range(load_size):
        for j in range(bus_size):
            mismatch[count] += load_info[i,2]*bus_info[j,4]*(g_matrix[int(load_info[i,0]),j]*np.sin(bus_info[i,5]-bus_info[j,5])-b_matrix[int(load_info[i,0]),j]*np.cos(bus_info[i,5]-bus_info[j,5]))
        mismatch[count] += bus_info[int(load_info[i,0]),3]
        if abs(mismatch[count]) > abs(max_mismatch[1, 0]):
            max_mismatch[1, 0] = abs(mismatch[count])
            max_mismatch[1, 1] = int(load_info[i,0]) + 1
        count += 1
    ws2.cell(row=iteration+2,column=1).value = iteration
    ws2.cell(row=iteration+2,column=2).value = max_mismatch[0,0]
    ws2.cell(row=iteration+2,column=3).value = max_mismatch[0,1]
    ws2.cell(row=iteration+2,column=4).value = max_mismatch[1,0]
    ws2.cell(row=iteration+2,column=5).value = max_mismatch[1,1]
    return mismatch, max_mismatch


# Make the Jacobian
def calc_jacobian(bus_info, load_info, g_matrix, b_matrix, bus_size, load_size):
    j_size = bus_size - 1 + load_size

    # dP/dTheta
    j11 = calc_j11(bus_info, load_info, g_matrix, b_matrix, bus_size, load_size)
    j11 = np.pad(j11, ((0, load_size), (0, load_size)), mode='constant', constant_values=(0, 0))

    # dP/dV
    j12 = calc_j12(bus_info, load_info, g_matrix, b_matrix, bus_size, load_size)
    j12 = np.pad(j12, ((0, load_size), (bus_size-1, 0)), mode='constant', constant_values=(0, 0))

    # dQ/dTheta
    j21 = calc_j21(bus_info, load_info, g_matrix, b_matrix, bus_size, load_size)
    j21 = np.pad(j21, ((bus_size-1, 0), (0, load_size)), mode='constant', constant_values=(0, 0))

    # dQ/dV
    j22 = calc_j22(bus_info, load_info, g_matrix, b_matrix, bus_size, load_size)
    j22 = np.pad(j22, ((bus_size-1, 0), (bus_size-1, 0)), mode='constant', constant_values=(0, 0))

    jacobian = j11 + j12 + j21 + j22
    return jacobian


# Constructs the J11 matrix
def calc_j11(bus_info, load_info, g_matrix, b_matrix, bus_size, load_size):
    j11 = np.zeros((bus_size - 1, bus_size - 1))
    for i in range(bus_size-1):
        for j in range(bus_size-1):
            if i == j:  # same bus
                term1 = -1 * (bus_info[i+1, 3] + bus_info[i+1, 1])
                term2 = b_matrix[i+1,i+1]*bus_info[i+1,4]**2
                j11[i,j] = term1 - term2
            else:
                term1 = g_matrix[i+1,j+1]*np.sin(bus_info[i+1,5]-bus_info[j+1,5])
                term2 = b_matrix[i+1,j+1]*np.cos(bus_info[i+1,5]-bus_info[j+1,5])
                j11[i,j] = bus_info[i+1,4]*bus_info[j+1,4]*(term1-term2)
    return j11


# Constructs the J12 matrix
def calc_j12(bus_info, load_info, g_matrix, b_matrix, bus_size, load_size):
    j12 = np.zeros((bus_size - 1, load_size))
    for i in range(bus_size - 1):
        for j in range(load_size):
            if i+1 == int(load_info[j, 0]):  # same bus
                term1 = (bus_info[i+1,0] - bus_info[i+1,2]) / bus_info[i+1,4]
                term2 = g_matrix[int(load_info[j, 0]),int(load_info[j, 0])] * bus_info[i+1,4]
                j12[i,j] = term1 + term2
            else:
                term1 = g_matrix[i+1,int(load_info[j, 0])]*np.cos(bus_info[i+1,5]-bus_info[int(load_info[j, 0]),5])
                term2 = b_matrix[i+1,int(load_info[j, 0])]*np.sin(bus_info[i+1,5]-bus_info[int(load_info[j, 0]),5])
                j12[i,j] = bus_info[i+1, 4]*(term1-term2)
    return j12


# Constructs the J21 matrix
def calc_j21(bus_info, load_info, g_matrix, b_matrix, bus_size, load_size):
    j21 = np.zeros((load_size, bus_size - 1))
    for i in range(load_size):
        for j in range(bus_size-1):
            if int(load_info[i,0]) == j+1:  # same bus
                term1 = bus_info[j+1,2] - bus_info[j+1,0]
                term2 = g_matrix[j+1,j+1]*bus_info[j+1,4]**2
                j21[i,j] = term1 - term2
            else:
                term1 = g_matrix[int(load_info[i,0]),j+1]*np.cos(bus_info[int(load_info[i,0]),5]-bus_info[j+1,5])
                term2 = b_matrix[int(load_info[i,0]),j+1]*np.sin(bus_info[int(load_info[i,0]),5]-bus_info[j+1,5])
                j21[i, j] = -1*bus_info[int(load_info[i,0]),4]*bus_info[j+1,4]*(term1+term2)
    return j21


# Constructs the J22 matrix
def calc_j22(bus_info, load_info, g_matrix, b_matrix, bus_size, load_size):
    j22 = np.zeros((load_size, load_size))
    for i in range(load_size):
        for j in range(load_size):
            bus1 = int(load_info[i,0])
            bus2 = int(load_info[j,0])
            if bus1==bus2:  # same bus
                term1 = (bus_info[bus1,3]+bus_info[bus1,1]) / bus_info[bus1,4]
                term2 = b_matrix[bus1,bus1]*bus_info[bus1,4]
                j22[i,j] = term1 - term2
            else:
                term1 = g_matrix[bus1,bus2]*np.sin(bus_info[bus1,5]-bus_info[bus2,5])
                term2 = b_matrix[bus1,bus2]*np.cos(bus_info[bus1,5]-bus_info[bus2,5])
                j22[i,j] = bus_info[bus1,4]*(term1-term2)
    return j22


main()
